from io import BytesIO

from openpyxl import load_workbook

from backend.schemas.reference_title import ReferenceTitleCreate


EXPECTED_HEADERS = {
    "titulo_investigacion": {
        "titulo_investigacion",
        "titulo investigacion",
        "titulo",
        "título_investigacion",
        "título investigación",
    },
    "linea_investigacion": {
        "linea_investigacion",
        "linea investigacion",
        "línea_investigacion",
        "línea investigación",
    },
    "sub_linea": {
        "sub_linea",
        "sub linea",
        "sublinea",
        "sub-línea",
        "sub línea",
    },
    "authors": {
        "authors",
        "author",
        "autores",
        "autor",
    },
    "status": {
        "status",
        "estado",
    },
}

REQUIRED_FIELDS = {"titulo_investigacion", "linea_investigacion", "sub_linea"}


def normalize_header(value: str | None) -> str:
    return (value or "").strip().lower().replace("-", " ").replace("_", " ")


def resolve_column_map(headers: list[str]) -> dict[str, int]:
    normalized_headers = [normalize_header(header) for header in headers]
    column_map: dict[str, int] = {}

    for field_name, aliases in EXPECTED_HEADERS.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        for index, header in enumerate(normalized_headers):
            if header in normalized_aliases:
                column_map[field_name] = index
                break

    missing_fields = [field for field in REQUIRED_FIELDS if field not in column_map]
    if missing_fields:
        raise ValueError(
            "El Excel debe incluir las columnas titulo_investigacion, linea_investigacion y sub_linea"
        )

    return column_map


def parse_reference_titles_excel(file_bytes: bytes) -> list[ReferenceTitleCreate]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("El archivo Excel no contiene filas")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    column_map = resolve_column_map(headers)
    items: list[ReferenceTitleCreate] = []

    for row_index, row in enumerate(rows[1:], start=2):
        titulo = str(row[column_map["titulo_investigacion"]]).strip() if row[column_map["titulo_investigacion"]] is not None else ""
        linea = str(row[column_map["linea_investigacion"]]).strip() if row[column_map["linea_investigacion"]] is not None else ""
        sub_linea = str(row[column_map["sub_linea"]]).strip() if row[column_map["sub_linea"]] is not None else ""
        authors_index = column_map.get("authors")
        status_index = column_map.get("status")

        authors = str(row[authors_index]).strip() if authors_index is not None and row[authors_index] is not None else ""
        status = str(row[status_index]).strip() if status_index is not None and row[status_index] is not None else "APROVADO"


        if not titulo and not linea and not sub_linea:
            continue

        if not titulo or not linea or not sub_linea:
            raise ValueError(f"La fila {row_index} tiene columnas obligatorias vacias")

        items.append(
            ReferenceTitleCreate(
                titulo_investigacion=titulo,
                linea_investigacion=linea,
                sub_linea=sub_linea,
                authors=authors,
                status=status or "APROVADO",
            )
        )

    if not items:
        raise ValueError("El archivo Excel no contiene registros validos")

    return items
